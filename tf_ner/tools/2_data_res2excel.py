import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

res_file = open('../models/lstm_crf/results_dde/score/testa.preds.txt', encoding='utf-8')
wb = Workbook()
wb.create_sheet('分类准确率，召回率，f1值')
ws = wb['Sheet']
ws1 = wb['分类准确率，召回率，f1值']
data = [[0, 0, 0, 0] for _ in range(10)]
red_fill = PatternFill("solid", fgColor="CCCCFF")  # 单元格填充颜色
line_count = 32940


lines = []

for i in range(line_count):
    line = res_file.readline().split()
    lines.append(line)
    word_tags = line
    if len(word_tags) < 3:
        continue
    else:
        if 'ROCK' in word_tags[1]:
            data[0][0] += 1
            if word_tags[1] == word_tags[2]:
                data[0][1] += 1
            elif word_tags[2] != 'O':
                data[0][2] += 1
            else:
                data[0][3] += 1
        if 'TECT' in word_tags[1]:
            data[1][0] += 1
            if word_tags[1] == word_tags[2]:
                data[1][1] += 1
            elif word_tags[2] != 'O':
                data[1][2] += 1
            else:
                data[1][3] += 1
        if 'ALTE' in word_tags[1]:
            data[2][0] += 1
            if word_tags[1] == word_tags[2]:
                data[2][1] += 1
            elif word_tags[2] != 'O':
                data[2][2] += 1
            else:
                data[2][3] += 1
        if 'PHYS' in word_tags[1]:
            data[3][0] += 1
            if word_tags[1] == word_tags[2]:
                data[3][1] += 1
            elif word_tags[2] != 'O':
                data[3][2] += 1
            else:
                data[3][3] += 1
        if 'CHEM' in word_tags[1]:
            data[4][0] += 1
            if word_tags[1] == word_tags[2]:
                data[4][1] += 1
            elif word_tags[2] != 'O':
                data[4][2] += 1
            else:
                data[4][3] += 1
        if 'CHRO' in word_tags[1]:
            data[5][0] += 1
            if word_tags[1] == word_tags[2]:
                data[5][1] += 1
            elif word_tags[2] != 'O':
                data[5][2] += 1
            else:
                data[5][3] += 1
        if 'MINE' in word_tags[1]:
            data[6][0] += 1
            if word_tags[1] == word_tags[2]:
                data[6][1] += 1
            elif word_tags[2] != 'O':
                data[6][2] += 1
            else:
                data[6][3] += 1
        if 'DEPO' in word_tags[1]:
            data[7][0] += 1
            if word_tags[1] == word_tags[2]:
                data[7][1] += 1
            elif word_tags[2] != 'O':
                data[7][2] += 1
            else:
                data[7][3] += 1
        if 'DEEP' in word_tags[1]:
            data[8][0] += 1
            if word_tags[1] == word_tags[2]:
                data[8][1] += 1
            elif word_tags[2] != 'O':
                data[8][2] += 1
            else:
                data[8][3] += 1
        if 'ELEM' in word_tags[1]:
            data[9][0] += 1
            if word_tags[1] == word_tags[2]:
                data[9][1] += 1
            elif word_tags[2] != 'O':
                data[9][2] += 1
            else:
                data[9][3] += 1


row = 1
col = 1
for i in range(len(lines)):
    sp = lines[i]
    print(sp)
    if len(sp) > 1:
        ws.cell(row, col).value = sp[0]
        ws.cell(row+1, col).value = sp[1]
        ws.cell(row+2, col).value = sp[2]
        if sp[1] != sp[2]:
            ws.cell(row+2, col).fill = red_fill
        col += 1
    else:
        row += 5
        col = 1

ws1.cell(1, 1).value = '类别'
ws1.cell(1, 2).value = '准确率'
ws1.cell(1, 3).value = '召回率'
ws1.cell(1, 4).value = 'f1 值'
ws1.cell(2, 1).value = 'ROCK'
ws1.cell(3, 1).value = 'TECT'
ws1.cell(4, 1).value = 'ALTE'
ws1.cell(5, 1).value = 'PHYS'
ws1.cell(6, 1).value = 'CHEM'
ws1.cell(7, 1).value = 'CHRO'
ws1.cell(8, 1).value = 'MINE'
ws1.cell(9, 1).value = 'DEPO'
ws1.cell(10, 1).value = 'DEEP'
ws1.cell(11, 1).value = 'ELEM'

for i in range(10):
    if data[i][1] == 0:
        ws1.cell(i+2, 2).value = 0
        ws1.cell(i+2, 3).value = 0
        ws1.cell(i+2, 4).value = 0
    else:
        ws1.cell(i+2, 2).value = data[i][1] / (data[i][1] + data[i][2])
        p = data[i][1] / (data[i][1] + data[i][2])
        ws1.cell(i+2, 3).value = data[i][1] / (data[i][1] + data[i][3])
        r = data[i][1] / (data[i][1] + data[i][3])
        ws1.cell(i+2, 4).value = 2*p*r / (p+r)

wb.save('./out_data/res_excel.xlsx')
